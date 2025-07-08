import pandas as pd
import openpyxl
from typing import Dict, List, Optional, Union
import os

def extract_text_from_excel(filepath: str, include_formulas: bool = False, 
                           include_metadata: bool = True, max_rows_per_sheet: int = 10000) -> str:
    """
    Enhanced Excel text extraction with better formatting and error handling.
    
    Args:
        filepath (str): Path to Excel file
        include_formulas (bool): Whether to extract formulas
        include_metadata (bool): Whether to include sheet metadata
        max_rows_per_sheet (int): Maximum rows to process per sheet
    
    Returns:
        str: Extracted text content
    """
    if not os.path.exists(filepath):
        print(f"❌ Excel file not found: {filepath}")
        return ""
    
    try:
        xl = pd.ExcelFile(filepath)
        content = []
        
        print(f"📊 Processing Excel file: {os.path.basename(filepath)}")
        print(f"📋 Found {len(xl.sheet_names)} sheets")
        
        for sheet_idx, sheet_name in enumerate(xl.sheet_names, 1):
            try:
                print(f"   Processing sheet {sheet_idx}/{len(xl.sheet_names)}: {sheet_name}")
                
                # Read sheet with error handling
                df = xl.parse(sheet_name, nrows=max_rows_per_sheet)
                
                if df.empty:
                    print(f"   ⚠️ Sheet '{sheet_name}' is empty, skipping")
                    continue
                
                # Process the sheet
                sheet_content = process_sheet(df, sheet_name, include_metadata)
                
                if sheet_content.strip():
                    content.append(sheet_content)
                    print(f"   ✅ Extracted {len(sheet_content)} characters from '{sheet_name}'")
                else:
                    print(f"   ⚠️ No content extracted from '{sheet_name}'")
                
            except Exception as e:
                print(f"   ❌ Error processing sheet '{sheet_name}': {e}")
                continue
        
        if not content:
            print("⚠️ No content extracted from any sheets")
            return ""
        
        final_content = "\n\n" + "="*50 + "\n\n".join(content)
        print(f"✅ Total extracted content: {len(final_content)} characters")
        
        return final_content
        
    except Exception as e:
        print(f"❌ Excel parse error: {e}")
        return ""

def process_sheet(df: pd.DataFrame, sheet_name: str, include_metadata: bool = True) -> str:
    """
    Process a single Excel sheet into readable text format.
    
    Args:
        df (pd.DataFrame): Sheet data
        sheet_name (str): Name of the sheet
        include_metadata (bool): Whether to include metadata
    
    Returns:
        str: Formatted sheet content
    """
    content_parts = []
    
    # Add sheet header
    content_parts.append(f"SHEET: {sheet_name}")
    content_parts.append("=" * (len(sheet_name) + 7))
    
    # Add metadata if requested
    if include_metadata:
        metadata = get_sheet_metadata(df)
        if metadata:
            content_parts.append("METADATA:")
            for key, value in metadata.items():
                content_parts.append(f"  {key}: {value}")
            content_parts.append("")
    
    # Clean the dataframe
    df_cleaned = clean_dataframe(df)
    
    if df_cleaned.empty:
        content_parts.append("(Empty sheet)")
        return "\n".join(content_parts)
    
    # Format the data
    formatted_data = format_dataframe_as_text(df_cleaned)
    content_parts.extend(formatted_data)
    
    return "\n".join(content_parts)

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Clean the dataframe by removing empty rows/columns and handling NaN values.
    
    Args:
        df (pd.DataFrame): Input dataframe
    
    Returns:
        pd.DataFrame: Cleaned dataframe
    """
    # Make a copy to avoid modifying original
    df_clean = df.copy()
    
    # Remove completely empty rows and columns
    df_clean = df_clean.dropna(how='all').dropna(axis=1, how='all')
    
    # Fill remaining NaN values with empty strings
    df_clean = df_clean.fillna("")
    
    # Convert all values to strings and strip whitespace
    for col in df_clean.columns:
        df_clean[col] = df_clean[col].astype(str).str.strip()
    
    # Remove rows where all values are empty strings
    df_clean = df_clean[~(df_clean == "").all(axis=1)]
    
    return df_clean

def format_dataframe_as_text(df: pd.DataFrame) -> List[str]:
    """
    Format dataframe as readable text with proper alignment.
    
    Args:
        df (pd.DataFrame): Dataframe to format
    
    Returns:
        List[str]: Formatted text lines
    """
    if df.empty:
        return ["(No data)"]
    
    formatted_lines = []
    
    # Check if first row looks like headers
    has_headers = detect_headers(df)
    
    if has_headers:
        # Use first row as headers
        headers = df.iloc[0].tolist()
        data_rows = df.iloc[1:]
        
        # Add headers
        formatted_lines.append("HEADERS:")
        formatted_lines.append(" | ".join(str(h) for h in headers))
        formatted_lines.append("-" * (len(" | ".join(str(h) for h in headers))))
        
        # Add data rows
        if not data_rows.empty:
            formatted_lines.append("DATA:")
            for idx, row in data_rows.iterrows():
                row_text = " | ".join(str(val) for val in row.tolist())
                if row_text.strip():  # Only add non-empty rows
                    formatted_lines.append(row_text)
    else:
        # Treat all rows as data
        formatted_lines.append("DATA:")
        for idx, row in df.iterrows():
            row_text = " | ".join(str(val) for val in row.tolist())
            if row_text.strip():  # Only add non-empty rows
                formatted_lines.append(row_text)
    
    return formatted_lines

def detect_headers(df: pd.DataFrame) -> bool:
    """
    Detect if the first row contains headers.
    
    Args:
        df (pd.DataFrame): Dataframe to analyze
    
    Returns:
        bool: True if first row appears to be headers
    """
    if df.empty or len(df) < 2:
        return False
    
    first_row = df.iloc[0]
    second_row = df.iloc[1]
    
    # Check if first row has text while second row has numbers/dates
    first_row_types = [type(val).__name__ for val in first_row]
    second_row_types = [type(val).__name__ for val in second_row]
    
    # Simple heuristic: if first row is mostly strings and second row has numbers
    first_str_count = sum(1 for t in first_row_types if 'str' in t.lower())
    second_num_count = sum(1 for val in second_row if str(val).replace('.', '').replace('-', '').isdigit())
    
    return first_str_count >= len(first_row) * 0.7 and second_num_count >= len(second_row) * 0.3

def get_sheet_metadata(df: pd.DataFrame) -> Dict[str, Union[str, int]]:
    """
    Extract metadata from the sheet.
    
    Args:
        df (pd.DataFrame): Dataframe to analyze
    
    Returns:
        Dict: Metadata information
    """
    if df.empty:
        return {"Status": "Empty sheet"}
    
    metadata = {
        "Rows": len(df),
        "Columns": len(df.columns),
        "Non-empty cells": df.notna().sum().sum(),
        "Data types": ", ".join(df.dtypes.astype(str).unique())
    }
    
    # Check for potential data patterns
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
    if numeric_cols:
        metadata["Numeric columns"] = len(numeric_cols)
    
    date_cols = df.select_dtypes(include=['datetime']).columns.tolist()
    if date_cols:
        metadata["Date columns"] = len(date_cols)
    
    return metadata

def extract_text_from_excel_advanced(filepath: str, extract_formulas: bool = False,
                                    extract_comments: bool = False) -> str:
    """
    Advanced Excel extraction using openpyxl for formulas and comments.
    
    Args:
        filepath (str): Path to Excel file
        extract_formulas (bool): Whether to extract formulas
        extract_comments (bool): Whether to extract comments
    
    Returns:
        str: Advanced extracted content
    """
    try:
        workbook = openpyxl.load_workbook(filepath, data_only=False)
        content = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_content = []
            
            sheet_content.append(f"SHEET: {sheet_name}")
            sheet_content.append("=" * (len(sheet_name) + 7))
            
            # Extract cell values, formulas, and comments
            for row in sheet.iter_rows():
                row_data = []
                row_formulas = []
                row_comments = []
                
                for cell in row:
                    # Cell value
                    value = cell.value if cell.value is not None else ""
                    row_data.append(str(value))
                    
                    # Cell formula
                    if extract_formulas and hasattr(cell, 'formula') and cell.formula:
                        row_formulas.append(f"{cell.coordinate}: {cell.formula}")
                    
                    # Cell comment
                    if extract_comments and cell.comment:
                        row_comments.append(f"{cell.coordinate}: {cell.comment.text}")
                
                # Add row data if not empty
                row_text = " | ".join(row_data)
                if row_text.strip():
                    sheet_content.append(row_text)
                
                # Add formulas if any
                if row_formulas:
                    sheet_content.append(f"FORMULAS: {'; '.join(row_formulas)}")
                
                # Add comments if any
                if row_comments:
                    sheet_content.append(f"COMMENTS: {'; '.join(row_comments)}")
            
            if len(sheet_content) > 2:  # More than just header
                content.append("\n".join(sheet_content))
        
        return "\n\n" + "="*50 + "\n\n".join(content)
        
    except Exception as e:
        print(f"❌ Advanced Excel extraction error: {e}")
        # Fallback to basic extraction
        return extract_text_from_excel(filepath)

def extract_specific_sheets(filepath: str, sheet_names: List[str]) -> str:
    """
    Extract text from specific sheets only.
    
    Args:
        filepath (str): Path to Excel file
        sheet_names (List[str]): List of sheet names to extract
    
    Returns:
        str: Extracted content from specified sheets
    """
    try:
        xl = pd.ExcelFile(filepath)
        content = []
        
        for sheet_name in sheet_names:
            if sheet_name in xl.sheet_names:
                df = xl.parse(sheet_name)
                sheet_content = process_sheet(df, sheet_name)
                if sheet_content.strip():
                    content.append(sheet_content)
            else:
                print(f"⚠️ Sheet '{sheet_name}' not found in file")
        
        return "\n\n" + "="*50 + "\n\n".join(content)
        
    except Exception as e:
        print(f"❌ Specific sheet extraction error: {e}")
        return ""

def get_excel_info(filepath: str) -> Dict:
    """
    Get information about an Excel file without extracting content.
    
    Args:
        filepath (str): Path to Excel file
    
    Returns:
        Dict: File information
    """
    try:
        xl = pd.ExcelFile(filepath)
        info = {
            "file_name": os.path.basename(filepath),
            "file_size": os.path.getsize(filepath),
            "sheet_count": len(xl.sheet_names),
            "sheet_names": xl.sheet_names,
            "sheets_info": {}
        }
        
        for sheet_name in xl.sheet_names:
            try:
                df = xl.parse(sheet_name, nrows=1)  # Just get shape info
                full_df = xl.parse(sheet_name)
                info["sheets_info"][sheet_name] = {
                    "rows": len(full_df),
                    "columns": len(full_df.columns),
                    "has_data": not full_df.empty
                }
            except:
                info["sheets_info"][sheet_name] = {"error": "Could not read sheet"}
        
        return info
        
    except Exception as e:
        return {"error": f"Could not read Excel file: {e}"}

# Test and utility functions
def test_excel_parser(filepath: str):
    """Test the Excel parser with a file."""
    print(f"🧪 Testing Excel parser with: {filepath}")
    
    # Basic extraction
    print("\n📊 Basic extraction:")
    basic_content = extract_text_from_excel(filepath)
    print(f"Extracted {len(basic_content)} characters")
    
    # File info
    print("\n📋 File information:")
    info = get_excel_info(filepath)
    print(f"Sheets: {info.get('sheet_count', 'unknown')}")
    print(f"Sheet names: {info.get('sheet_names', [])}")
    
    # Advanced extraction
    print("\n🔧 Advanced extraction:")
    advanced_content = extract_text_from_excel_advanced(filepath, extract_formulas=True)
    print(f"Advanced extracted {len(advanced_content)} characters")
    
    return basic_content

if __name__ == "__main__":
    # Example usage
    test_file = "sample.xlsx"  # Replace with your test file
    if os.path.exists(test_file):
        result = test_excel_parser(test_file)
        print("\n📄 Sample output:")
        print(result[:500] + "..." if len(result) > 500 else result)
    else:
        print(f"Test file {test_file} not found. Create a sample Excel file to test.")

